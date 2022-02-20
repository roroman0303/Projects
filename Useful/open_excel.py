import openpyxl

# открыть лист в таблице Excel и получить её в виде двумерного списка
def open_table(filename='korean_doc.xlsx', sheetname='database(en)'):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    table = list()
    for row in range(0, sheet.max_row+1):
        table.append([])
        for col in range(0, sheet.max_column+1):
            current_cell = sheet.cell(row=row + 1, column=col + 1).value
            if str(current_cell).isspace() != True and current_cell != '': # проверяем на пустоту
                table[row].append(current_cell) # +1 потому что cell-ы с 1
            else: # если пустая, то пусть None
                table[row].append(None)
    return table
