import telebot
import os
from openpyxl import load_workbook

with open('trusted_users.txt', 'r') as file:
    trusted_users = file.read().splitlines()

with open('telegram_token.txt', 'r') as file:
    lines = file.read().splitlines()
    token = lines[0]
bot = telebot.TeleBot(token)

with open('sheet_name.txt', 'r') as file:
    sheet_name = file.read().splitlines()[0]

table = load_workbook('list.xlsx')
sheet = table[sheet_name]

keyboard = telebot.types.ReplyKeyboardMarkup
hi_counter = 0

print('Бот включен и готов к работе.')
print('t.me/Maksim_Excel_Bot')

@bot.message_handler(commands=['start']) # что бот будет делать на команды start
def welcome(message):
    print('Пользователь {} вызвал команду /start'.format(message.from_user.first_name))
    menu = keyboard(True, True)
    menu.row('Открой таблицу')
    bot.send_message(message.chat.id, 'Привет, {}. Я могу открыть Excel таблицу.'.format(message.from_user.first_name), reply_markup=menu)

@bot.message_handler(commands=['help']) # что бот будет делать на команды help
def welcome(message):
    print('Пользователь {} вызвал команду /help'.format(message.from_user.first_name))
    bot.send_message(message.chat.id, 'Спроси Романа.')


@bot.message_handler(content_types=['text']) # что бот будет делать на команды start и help
def answer_to_text(message):
    global trusted_users
    print('Пользователь {}({}) написал сообщение: {}'.format(message.from_user.first_name, message.from_user.id, message.text))
    if 'привет' in message.text.lower():
        global hi_counter
        hi_counter += 1
        menu = keyboard(True, True)
        menu.row('Открой таблицу')
        bot.send_message(message.from_user.id, 'Привет-привет!')
        bot.send_message(message.from_user.id, 'С прошлого рестарта со мной поздаровались уже {} раз!'.format(hi_counter), reply_markup=menu)

    elif 'Открой таблицу' == message.text or 'Продолжить' == message.text :
        if str(message.from_user.id) not in trusted_users:
            bot.send_message(message.from_user.id, 'Введите пароль.')
        else:
            bot.send_message(message.from_user.id, 'Введите номер.')

    elif 'пароль' in message.text.lower():
        if str(message.from_user.id) not in trusted_users:
            trusted_users.append(message.from_user.id)
            with open('trusted_users.txt', 'w') as file:
                for line in trusted_users:
                    file.write(str(line) + '\n')
        menu = keyboard(True, True)
        menu.row('Открой таблицу')
        bot.send_message(message.from_user.id, 'Так бы сразу. Добро пожаловать!', reply_markup=menu)

    elif message.text.isdigit():
        if str(message.from_user.id) not in trusted_users:
            bot.send_message(message.from_user.id, 'Введите пароль.')
        else:
            global sheet
            result = sheet.cell(row=int(message.text), column=2).value
            menu = keyboard(True, True)
            menu.row('Продолжить')
            bot.send_message(message.from_user.id, result, reply_markup=menu)
    else:
        menu = keyboard(True, True)
        menu.row('Открой таблицу')
        bot.send_message(message.from_user.id, 'Не понял :(', reply_markup=menu)

# ===========================================

bot.polling(none_stop=True)