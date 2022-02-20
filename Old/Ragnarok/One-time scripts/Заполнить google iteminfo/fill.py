import Ragnarok.module.google as google
import Ragnarok.module.files as files
import Ragnarok.module.find as find
import re
import openpyxl

wb = openpyxl.load_workbook('desc.xlsx')
sheet = wb['database']

worksheet = google.open_sheet('1HchGulIn5IHlQYJCg9y4XI-IkLqzUhW05rnVsizUoEw', 1)
worksheet.update_cell(1, 2, 'DB name') # строка столбец что

iteminfo_ru = files.open_lines(files.get_path('ropru', 'iteminfo.lua'), codirovka='Windows-1251', errors='ignore')
iteminfo_en = files.open_lines(files.get_path('ropeu', 'iteminfo.lua'), codirovka='EUC-KR', errors='ignore')

print('открываю desc.xlsx')
old_table = files.open_table('desc.xlsx', 'database')
for row in old_table:
    if row[0] is None:
        current_row = old_table.index(row) + 1
        break

ids = list()
print('Ищу id в iteminfo ru')
for line in iteminfo_ru:
    if re.search(r'\s*\[(.*)\]\s*=', line):
        try:
            id = int(re.search(r'^\s*\[(.*)\]\s*=', line).group(1).strip())
        except:
            continue
        if id not in ids: ids.append(id)
print('Ищу id в iteminfo en')
for line in iteminfo_en:
    if re.search(r'\s*\[(.*)\]\s*=', line):
        try:
            id = int(re.search(r'^\s*\[(.*)\]\s*=', line).group(1).strip())
        except: continue
        if id not in ids: ids.append(id)

for id in ids:
    print(id)
    done = files.open_lines('done.txt')
    if str(id) in done: continue
    name_ru = ''
    name_en = ''
    desc_ru = ''
    desc_en = ''
    info_ru = find.iteminfo('ropru', id)
    info_en = find.iteminfo('ropeu', id)
    if info_ru is not None:
        for item_line in info_ru.splitlines():
            if 'identifiedDisplayName' in item_line and 'unidentifiedDisplayName' not in item_line:
                name_ru = item_line.split("\"")[1]
                break
        try:
            desc_ru = info_ru.split("unidentifiedDescriptionName")[1]
            desc_ru = desc_ru.split("identifiedDescriptionName")[1]
            desc_ru = desc_ru.split("\"", 1)[1]
            desc_ru = desc_ru.rsplit("\"", 1)[0]
            desc_ru = desc_ru.replace("\"","")
            desc_ru_lines = list()
            for desc_line in desc_ru.splitlines():
                if desc_line[-1] == ',': desc_line = desc_line[:-1]
                desc_ru_lines.append(desc_line.strip())
            desc_ru = '\n'.join(desc_ru_lines)
        except:
            desc_ru = ''
    if info_en is not None:
        for item_line in info_en.splitlines():
            if 'identifiedDisplayName' in item_line and 'unidentifiedDisplayName' not in item_line:
                name_en = item_line.split("\"")[1]
                break
        try:
            desc_en = info_en.split("unidentifiedDescriptionName")[1]
            desc_en = desc_en.split("identifiedDescriptionName")[1]
            desc_en = desc_en.split("\"", 1)[1]
            desc_en = desc_en.rsplit("\"", 1)[0]
            desc_en = desc_en.replace("\"", "")
            desc_en_lines = list()
            for desc_line in desc_en.splitlines():
                if len(desc_line) == 0: continue
                if desc_line[-1] == ',': desc_line = desc_line[:-1]
                desc_en_lines.append(desc_line.strip())
            desc_en = '\n'.join(desc_en_lines)
        except:
            desc_en = ''

    print(name_ru)
    print(name_en)
    print(desc_ru)
    print(desc_en)
    sheet.cell(column=1, row=current_row, value=id)
    sheet.cell(column=3, row=current_row, value=name_ru)
    sheet.cell(column=4, row=current_row, value=name_en)
    sheet.cell(column=5, row=current_row, value=desc_ru)
    sheet.cell(column=6, row=current_row, value=desc_en)
    current_row += 1
    wb.save(filename = 'desc.xlsx')
    done.append(str(id))
    files.write_lines(done, 'done.txt')