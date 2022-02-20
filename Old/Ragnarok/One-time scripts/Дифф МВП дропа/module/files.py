import openpyxl
import json
import os
import subprocess
import shutil

files_paths = {'special.sc' : '/zone/itemdata/',
               'script.sql' : '/dbscripts/',
               'Importantitem.txt' : '/zone/',
               'CashItemList.txt': '/zone/',
               'ItemMoveInfoV4.txt' : '/zone/',
               'ItemMoveInfoV5.txt' : '/zone/',
               'combiItem.sc' : '/zone/itemdata/',
               'BuffSpecial.sc' : '/zone/',
               'PackageItem.lua': '/zone/',
               'iteminfo.lua': '/data/gameFolder/',
               'itemslotcounttable.txt': '/data/',
               'num2itemresnametable.txt' : '/data/',
               'idnum2itemresnametable.txt' : '/data/',
               'accname.lua' : '/data/luafiles514/lua files/datainfo/',
               'accessoryid.lua' : '/data/luafiles514/lua files/datainfo/',
               'spriterobeid.lua' : '/data/luafiles514/lua files/datainfo/',
               'spriterobename.lua' : '/data/luafiles514/lua files/datainfo/',
               'tb_layer_priority.lua' : '/data/luafiles514/lua files/datainfo/',
               'ItemSummonList.txt': '/zone/',  # в последнюю очередь
               'Collection.lua' : '/zone/', # в последнюю очередь
               'NotDisappearAfterUsingItemList.lua' : '/zone/' # в последнюю очередь
               }

# открыть файл и получить его в виде списка строк
def open_lines(filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'r', encoding=codirovka, errors=errors) as file:
            lines = file.read().splitlines()
        return lines
    else: 
        with open(filename, 'r', encoding=codirovka) as file:
            lines = file.read().splitlines()
        return lines

# открыть файл и получить его в виде одной строки с текстом
def open_text(filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'r', encoding=codirovka, errors=errors) as file:
            text = file.read()
        return text
    else: 
        with open(filename, 'r', encoding=codirovka) as file:
            text = file.read()
        return text

# записать в файл список строк
def write_lines(lines, filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'w', encoding=codirovka, errors=errors) as file:
            for line in lines:
                file.write(line + '\n')
        return lines
    else: 
        with open(filename, 'w', encoding=codirovka) as file:
            for i in range(len(lines)):
                try:
                    file.write(lines[i] + '\n')
                except UnicodeEncodeError:
                    print("\033[31mВнимание\033[0m - при записи {} пропущена строка №{}: {}".format(filename, i, lines[i]))
                    file.write('// missing_line \n')
        return lines

# записать в файл строку с текстом
def write_text(text, filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'w', encoding=codirovka, errors=errors) as file:
            file.write(text)
    else:
        try:
            with open(filename, 'w', encoding=codirovka) as file:
                file.write(text)
        except UnicodeEncodeError:
            write_lines(text.splitlines(), filename, codirovka)

# открыть лист в таблице Excel и получить её в виде двумерного списка
def open_table(filename='korean_doc.xlsx', sheetname='database(en)'):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    table = list()
    for row in range(0, sheet.max_row+1):
        table.append([])
        for col in range(0, sheet.max_column+1):
            current_cell = sheet.cell(row=row + 1, column=col + 1).value
            if str(current_cell).isspace() != True and current_cell != '': # проверяем на пустоту
                table[row].append(current_cell) # +1 потому что cell-ы с 1
            else: # если пустая, то пусть None
                table[row].append(None)
    return table

# открыть файл в notepad++
def open_with_notepad(filenames):
    with open("paths.json", "r") as paths_json:
        notepad_path = json.load(paths_json)["notepad++"].replace('.exe', '')
    if isinstance(filenames, str):
        output = subprocess.check_output('\"{}\" \"{}\"'.format(notepad_path, filenames))
    else:
        output = subprocess.check_output('\"{}\" -multiInst -nosession \"{}\"'.format(notepad_path, filenames[0]))
        for filename in filenames[1:]:
            output = subprocess.check_output('\"{}\" \"{}\"'.format(notepad_path, filename))
    return output

# === получить расположение файла по региону из json-а ===
def get_path(region, file):
    with open("paths.json", "r") as paths_json:
        path = json.load(paths_json)[region]
    path += files_paths[file] + file
    return path

# === получить адрес google-таблицы из json-а ===
def get_url(id):
    with open("urls.json", "r") as paths_json:
        return json.load(paths_json)[id]

# === удалить папку с содержимым ===
def delete_folder(folder):
    shutil.rmtree(folder)

# === очистить папку ===
def clean_folder(folder):
    delete_folder(folder)
    os.mkdir(folder)




