import openpyxl
import json
import os
import subprocess
import shutil
from datetime import datetime
import time

ropru_files_paths = {'special.sc' : '/zone/itemdata/',
               'script.sql' : '/dbscripts/',
               'Importantitem.txt' : '/zone/',
               'CashItemList.txt': '/zone/',
               'SimpleCashShopScript.lua': '/zone/',
               'SilvervineExchange.sc' : '/zone/npcdata/trader/',
               'ItemMoveInfoV4.txt' : '/zone/',
               'ItemMoveInfoV5.txt' : '/zone/',
               'combiItem.sc' : '/zone/itemdata/',
               'BuffSpecial.sc' : '/zone/',
               'PackageItem.lua': '/zone/',
               'iteminfo.lua': '/data/gameFolder/',
               'itemslotcounttable.txt': '/data/',
               'num2itemresnametable.txt' : '/data/',
               'idnum2itemresnametable.txt' : '/data/',
               'accname.lua' : '/data/luafiles514/lua files/datainfo/',
               'accessoryid.lua' : '/data/luafiles514/lua files/datainfo/',
               'spriterobeid.lua' : '/data/luafiles514/lua files/datainfo/',
               'spriterobename.lua' : '/data/luafiles514/lua files/datainfo/',
               'tb_layer_priority.lua' : '/data/luafiles514/lua files/datainfo/',
               'ItemSummonList.txt': '/zone/',  # в последнюю очередь
               'Collection.lua' : '/zone/', # в последнюю очередь
               'NotDisappearAfterUsingItemList.lua' : '/zone/'}
ropeu_files_paths = {'special.sc' : '/zone/itemdata/',
               'script.sql' : '/dbscripts/',
               'Importantitem.txt' : '/zone/',
               'CashItemList.txt': '/zone/',
               'SimpleCashShopScript.lua': '/zone/',
               'SilvervineExchange.sc' : '/zone/npcdata/trader/',
               'ItemMoveInfoV4.txt' : '/zone/',
               'ItemMoveInfoV5.txt' : '/zone/',
               'combiItem.sc' : '/zone/itemdata/',
               'BuffSpecial.sc' : '/zone/',
               'PackageItem.lua': '/zone/',
               'iteminfo.lua': '/data/gameFolder/',
               'itemslotcounttable.txt': '/data/',
               'num2itemresnametable.txt' : '/data/',
               'idnum2itemresnametable.txt' : '/data/',
               'accname.lua' : '/data/luafiles514/lua files/datainfo/',
               'accessoryid.lua' : '/data/luafiles514/lua files/datainfo/',
               'spriterobeid.lua' : '/data/luafiles514/lua files/datainfo/',
               'spriterobename.lua' : '/data/luafiles514/lua files/datainfo/',
               'tb_layer_priority.lua' : '/data/luafiles514/lua files/datainfo/',
               'ItemSummonList.txt': '/zone/',  # в последнюю очередь
               'Collection.lua' : '/zone/', # в последнюю очередь
               'NotDisappearAfterUsingItemList.lua' : '/zone/'}
roeu_files_paths = {'special.sc' : '/zone/itemdata/',
               'script.sql' : '/dbscript/planned/',
               'Importantitem.txt' : '/zone/',
               'CashItemList.txt': '/zone/',
               'SimpleCashShopScript.lua': '/zone/',
               'SilvervineExchange.sc' : '/zone/npcdata/trader/',
               'ItemMoveInfoV4.txt' : '/zone/',
               'ItemMoveInfoV5.txt' : '/zone/',
               'combiItem.sc' : '/zone/itemdata/',
               'BuffSpecial.sc' : '/zone/',
               'PackageItem.lua': '/zone/',
               'iteminfo.lua': '/data/gameFolder/',
               'itemslotcounttable.txt': '/data/',
               'num2itemresnametable.txt' : '/data/',
               'idnum2itemresnametable.txt' : '/data/',
               'accname.lua' : '/data/luafiles514/lua files/datainfo/',
               'accessoryid.lua' : '/data/luafiles514/lua files/datainfo/',
               'spriterobeid.lua' : '/data/luafiles514/lua files/datainfo/',
               'spriterobename.lua' : '/data/luafiles514/lua files/datainfo/',
               'tb_layer_priority.lua' : '/data/luafiles514/lua files/datainfo/',
               'ItemSummonList.txt': '/zone/',  # в последнюю очередь
               'Collection.lua' : '/zone/', # в последнюю очередь
               'NotDisappearAfterUsingItemList.lua' : '/zone/'}
ru_codirovka = {'special.sc' : 'Windows-1251',
               'script.sql' : 'utf-8',
               'Importantitem.txt' : 'EUC-KR',
               'CashItemList.txt': 'EUC-KR',
               'SimpleCashShopScript.lua': 'utf-8',
               'SilvervineExchange.sc' : 'Windows-1251',
               'ItemMoveInfoV4.txt' : 'EUC-KR',
               'ItemMoveInfoV5.txt' : 'EUC-KR',
               'combiItem.sc' : 'utf-8',
               'BuffSpecial.sc' : 'EUC-KR',
               'PackageItem.lua': 'utf-8',
               'iteminfo.lua': 'Windows-1251',
               'itemslotcounttable.txt': 'EUC-KR',
               'num2itemresnametable.txt' : 'EUC-KR',
               'idnum2itemresnametable.txt' : 'EUC-KR',
               'accname.lua' : 'EUC-KR',
               'accessoryid.lua' : 'EUC-KR',
               'spriterobeid.lua' : 'EUC-KR',
               'spriterobename.lua' : 'EUC-KR',
               'tb_layer_priority.lua' : 'EUC-KR',
               'ItemSummonList.txt': 'EUC-KR',  # в последнюю очередь
               'Collection.lua' : 'EUC-KR', # в последнюю очередь
               'NotDisappearAfterUsingItemList.lua' : 'EUC-KR' # в последнюю очередь
               }
eu_codirovka = {'special.sc' : 'EUC-KR',
               'script.sql' : 'utf-8',
               'Importantitem.txt' : 'EUC-KR',
               'CashItemList.txt': 'EUC-KR',
               'SimpleCashShopScript.lua': 'utf-8',
               'SilvervineExchange.sc' : 'EUC-KR',
               'ItemMoveInfoV4.txt' : 'EUC-KR',
               'ItemMoveInfoV5.txt' : 'EUC-KR',
               'combiItem.sc' : 'utf-8',
               'BuffSpecial.sc' : 'EUC-KR',
               'PackageItem.lua': 'utf-8',
               'iteminfo.lua': 'EUC-KR',
               'itemslotcounttable.txt': 'EUC-KR',
               'num2itemresnametable.txt' : 'EUC-KR',
               'idnum2itemresnametable.txt' : 'EUC-KR',
               'accname.lua' : 'EUC-KR',
               'accessoryid.lua' : 'EUC-KR',
               'spriterobeid.lua' : 'EUC-KR',
               'spriterobename.lua' : 'EUC-KR',
               'tb_layer_priority.lua' : 'EUC-KR',
               'ItemSummonList.txt': 'EUC-KR',  # в последнюю очередь
               'Collection.lua' : 'EUC-KR', # в последнюю очередь
               'NotDisappearAfterUsingItemList.lua' : 'EUC-KR' # в последнюю очередь
               }

# открыть файл и получить его в виде списка строк
def open_lines(filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'r', encoding=codirovka, errors=errors) as file:
            lines = file.read().splitlines()
        return lines
    else:
        with open(filename, 'r', encoding=codirovka) as file:
            lines = file.read().splitlines()
        return lines

# открыть файл и получить его в виде одной строки с текстом
def open_text(filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'r', encoding=codirovka, errors=errors) as file:
            text = file.read()
        return text
    else:
        with open(filename, 'r', encoding=codirovka) as file:
            text = file.read()
        return text

# записать в файл список строк
def write_lines(lines, filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'w', encoding=codirovka, errors=errors) as file:
            for line in lines:
                file.write(line + '\n')
        return lines
    else:
        with open(filename, 'w', encoding=codirovka) as file:
            for i in range(len(lines)):
                try:
                    file.write(lines[i] + '\n')
                except UnicodeEncodeError:
                    if filename.startswith('history/'): pass
                    else: print("\033[31mВнимание\033[0m - при записи {} пропущена строка №{}: {}".format(filename, i+1, lines[i]))
                    file.write('// missing_line \n')
        return lines

# записать в файл строку с текстом
def write_text(text, filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'w', encoding=codirovka, errors=errors) as file:
            file.write(text)
    else:
        try:
            with open(filename, 'w', encoding=codirovka) as file:
                file.write(text)
        except UnicodeEncodeError:
            write_lines(text.splitlines(), filename, codirovka)

# открыть лист в таблице Excel и получить её в виде двумерного списка
def open_table(filename='korean_doc.xlsx', sheetname='database(en)'):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    table = list()
    for row in range(0, sheet.max_row+1):
        table.append([])
        for col in range(0, sheet.max_column+1):
            current_cell = sheet.cell(row=row + 1, column=col + 1).value
            if str(current_cell).isspace() != True and current_cell != '': # проверяем на пустоту
                table[row].append(current_cell) # +1 потому что cell-ы с 1
            else: # если пустая, то пусть None
                table[row].append(None)
    return table
