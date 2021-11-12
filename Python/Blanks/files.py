import openpyxl
import json
import os
import subprocess
import shutil
import time

# открыть файл и получить его в виде списка строк
def open_lines(filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'r', encoding=codirovka, errors=errors) as file:
            lines = file.read().splitlines()
        return lines
    else: 
        with open(filename, 'r', encoding=codirovka) as file:
            lines = file.read().splitlines()
        return lines

# открыть файл и получить его в виде одной строки с текстом
def open_text(filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'r', encoding=codirovka, errors=errors) as file:
            text = file.read()
        return text
    else: 
        with open(filename, 'r', encoding=codirovka) as file:
            text = file.read()
        return text

# записать в файл список строк
def write_lines(lines, filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'w', encoding=codirovka, errors=errors) as file:
            for line in lines:
                file.write(line + '\n')
        return lines
    else: 
        with open(filename, 'w', encoding=codirovka) as file:
            for i in range(len(lines)):
                try:
                    file.write(lines[i] + '\n')
                except UnicodeEncodeError:
                    if filename.startswith('history/'): pass
                    else: print("\033[31mВнимание\033[0m - при записи {} пропущена строка №{}: {}".format(filename, i+1, lines[i]))
                    file.write('// missing_line \n')
        return lines

# записать в файл строку с текстом
def write_text(text, filename, codirovka = 'utf-8', errors = ''):
    if errors != '':
        with open(filename, 'w', encoding=codirovka, errors=errors) as file:
            file.write(text)
    else:
        try:
            with open(filename, 'w', encoding=codirovka) as file:
                file.write(text)
        except UnicodeEncodeError:
            write_lines(text.splitlines(), filename, codirovka)

# открыть лист в таблице Excel и получить её в виде двумерного списка
def open_table(filename='doc.xlsx', sheetname='Лист 1'):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    table = list()
    for row in range(0, sheet.max_row+1):
        table.append([])
        for col in range(0, sheet.max_column+1):
            current_cell = sheet.cell(row=row + 1, column=col + 1).value
            if str(current_cell).isspace() != True and current_cell != '': # проверяем на пустоту
                table[row].append(current_cell)
            else: # если пустая, то пусть None
                table[row].append(None)
    return table

# открыть файл в notepad++
def open_with_notepad(filenames):
    with open("paths.json", "r") as paths_json:
        notepad_path = json.load(paths_json)["notepad++"].replace('.exe', '')
    output = subprocess.check_output('\"{}\" -multiInst -nosession \"{}\"'.format(notepad_path, filenames))
    return output

# получить адрес google-таблицы из json-а
def get_url(id):
    with open("urls.json", "r") as paths_json:
        return json.load(paths_json)[id]

# удалить папку с содержимым
def delete_folder(folder):
    try:
        shutil.rmtree(folder)
    except: pass

# очистить папку
def clean_folder(folder):
    delete_folder(folder)
    time.sleep(0.5)
    os.mkdir(folder)

# создать папку
def create_folder(folder):
    os.mkdir(folder)

# получить первое попавшееся имя таблицы xlsx
def get_xlsx_name():
    try:
        filenames = os.listdir('korean/')
    except:
        os.mkdir("korean")
        time.sleep(0.5)
        filenames = os.listdir('korean/')
    for filename in filenames:
        if '.xlsx' in filename:
            xlsx_filename = filename
            break
    try: return 'korean/'+xlsx_filename
    except: raise FileNotFoundError('Таблица Excel не найдена!')

